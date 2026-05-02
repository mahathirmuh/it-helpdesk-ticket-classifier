import openpyxl
from openpyxl.comments import Comment

OUTPUT_FILE = "results/cobacek_compare.xlsx"

NOTES_PREDICTIONS = {
    "subject":      "Judul / subjek tiket IT helpdesk",
    "description":  "Deskripsi lengkap tiket — input utama yang digunakan semua model",
    "answer":       "Jawaban atau solusi yang diberikan untuk tiket ini",
    "type":         "Tipe tiket (contoh: Request, Incident)",
    "priority":     "Prioritas AKTUAL (ground truth) dari tiket",
    "category":     "Kategori AKTUAL (ground truth) dari tiket",
    "svm_category": "Prediksi kategori oleh SVM (TF-IDF + LinearSVC, OOF K-Fold)",
    "svm_priority": "Prediksi prioritas oleh SVM (TF-IDF + LinearSVC, OOF K-Fold)",
    "rf_category":  "Prediksi kategori oleh Random Forest (TF-IDF, OOF K-Fold)",
    "rf_priority":  "Prediksi prioritas oleh Random Forest (TF-IDF, OOF K-Fold)",
    "lr_category":  "Prediksi kategori oleh Logistic Regression (TF-IDF, OOF K-Fold)",
    "lr_priority":  "Prediksi prioritas oleh Logistic Regression (TF-IDF, OOF K-Fold)",
    "svm_match":    "TRUE jika SVM benar di category DAN priority sekaligus",
    "rf_match":     "TRUE jika Random Forest benar di category DAN priority sekaligus",
    "lr_match":     "TRUE jika Logistic Regression benar di category DAN priority sekaligus",
    "bert_category": "Prediksi kategori oleh BERT (DistilBERT multilingual, OOF K-Fold)",
    "bert_priority": "Prediksi prioritas oleh BERT (DistilBERT multilingual, OOF K-Fold)",
    "bert_match":   "TRUE jika BERT benar di category DAN priority sekaligus",
}

NOTES_METRICS = {
    "approach":        "Nama model / pendekatan yang dibandingkan (SVM, RF, LR, BERT, Hybrid SVM)",
    "label":           "Label yang dievaluasi: 'category' atau 'priority'",
    "elapsed_seconds": "Total waktu training dan prediksi seluruh fold (dalam detik)",
    "accuracy":        "Akurasi = jumlah prediksi benar / total baris",
    "macro_precision": "Precision rata-rata antar kelas (tidak mempertimbangkan jumlah sampel per kelas)",
    "macro_recall":    "Recall rata-rata antar kelas (tidak mempertimbangkan jumlah sampel per kelas)",
    "macro_f1":        "F1-score rata-rata antar kelas (tidak mempertimbangkan jumlah sampel per kelas)",
    "weighted_f1":     "F1-score rata-rata tertimbang — kelas dengan sampel lebih banyak diberi bobot lebih besar",
    "samples":         "Jumlah total baris yang dievaluasi",
}

NOTES_SUMMARY = {
    "key":   "Nama parameter konfigurasi yang digunakan saat menjalankan notebook",
    "value": "Nilai dari parameter konfigurasi tersebut",
}

def add_header_comments(ws, notes_dict, hybrid_prefix=False):
    count = 0
    for cell in ws[1]:
        col_name = cell.value
        if col_name is None:
            continue
        note = notes_dict.get(col_name)
        if note is None and hybrid_prefix:
            if col_name.startswith("hybrid_svm_category_"):
                model_id = col_name[len("hybrid_svm_category_"):]
                note = f"Prediksi kategori Hybrid SVM: SVM dikoreksi GenAI ({model_id}) untuk baris mismatch"
            elif col_name.startswith("hybrid_svm_priority_"):
                model_id = col_name[len("hybrid_svm_priority_"):]
                note = f"Prediksi prioritas Hybrid SVM: SVM dikoreksi GenAI ({model_id}) untuk baris mismatch"
        if note:
            comment = Comment(note, "System")
            comment.width = 340
            comment.height = 60
            cell.comment = comment
            count += 1
    return count

wb = openpyxl.load_workbook(OUTPUT_FILE)

sheets = wb.sheetnames
print(f"Sheet ditemukan: {sheets}")

c1 = add_header_comments(wb["Predictions_Compare"], NOTES_PREDICTIONS, hybrid_prefix=True)
c2 = add_header_comments(wb["Metrics"],             NOTES_METRICS)
c3 = add_header_comments(wb["Summary"],             NOTES_SUMMARY)

wb.save(OUTPUT_FILE)
print(f"Selesai! Notes ditambahkan:")
print(f"  Predictions_Compare : {c1} kolom")
print(f"  Metrics             : {c2} kolom")
print(f"  Summary             : {c3} kolom")
