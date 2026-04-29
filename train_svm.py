import pandas as pd  # Library data processing utama
from openpyxl.comments import Comment  # Untuk menambah komentar pada sel Excel
from sklearn.feature_extraction.text import TfidfVectorizer  # Ekstraksi fitur teks TF-IDF
from sklearn.model_selection import train_test_split  # Membagi data train dan test
from sklearn.pipeline import Pipeline  # Menyusun alur pemrosesan dan model
from sklearn.svm import LinearSVC  # Model SVM linear untuk klasifikasi
from sklearn.metrics import classification_report, accuracy_score  # Metrik evaluasi klasifikasi


def train_and_evaluate(texts, labels):  # Fungsi untuk latih dan evaluasi model
    label_counts = labels.value_counts()  # Hitung jumlah data per kelas
    stratify_labels = labels if label_counts.min() >= 2 else None  # Gunakan stratify jika aman
    x_train, x_test, y_train, y_test = train_test_split(  # Split data
        texts, labels, test_size=0.2, random_state=42, stratify=stratify_labels  # Parameter split
    )
    pipeline = Pipeline(  # Rangkaian preprocessing + model
        [  # Definisi langkah pipeline
            ("tfidf", TfidfVectorizer(ngram_range=(1, 2))),  # TF-IDF unigram + bigram
            ("svm", LinearSVC()),  # SVM linear sebagai classifier
        ]  # Akhir langkah pipeline
    )  # Akhir pipeline
    pipeline.fit(x_train, y_train)  # Latih model pada data train
    y_pred = pipeline.predict(x_test)  # Prediksi pada data test
    report = classification_report(y_test, y_pred, zero_division=0)  # Laporan evaluasi
    report_dict = classification_report(  # Laporan evaluasi dalam format dict
        y_test, y_pred, zero_division=0, output_dict=True  # Parameter laporan
    )
    acc = accuracy_score(y_test, y_pred)  # Hitung akurasi
    return pipeline, report, report_dict, acc  # Kembalikan model dan metrik


def main():  # Fungsi utama program
    df = pd.read_excel("cobacek.xlsx")  # source data yang akan di predict
    if "description" not in df.columns:  # Validasi kolom description
        raise ValueError("Kolom 'description' tidak ditemukan.")  # Error jika tidak ada
    if "category" not in df.columns or "priority" not in df.columns:  # Validasi label
        raise ValueError("Kolom 'category' atau 'priority' tidak ditemukan.")  # Error jika tidak ada

    df["description"] = df["description"].fillna("").astype(str)  # Bersihkan teks description

    category_model, category_report, category_report_dict, category_acc = train_and_evaluate(  # Latih model category
        df["description"], df["category"]  # Input teks dan label category
    )
    priority_model, priority_report, priority_report_dict, priority_acc = train_and_evaluate(  # Latih model priority
        df["description"], df["priority"]  # Input teks dan label priority
    )

    print("Category Accuracy:", round(category_acc, 4))  # Tampilkan akurasi category
    print(category_report)  # Tampilkan laporan category
    print("Priority Accuracy:", round(priority_acc, 4))  # Tampilkan akurasi priority
    print(priority_report)  # Tampilkan laporan priority

    category_model.fit(df["description"], df["category"])  # Latih ulang category pada full data
    priority_model.fit(df["description"], df["priority"])  # Latih ulang priority pada full data

    df["predicted_category"] = category_model.predict(df["description"])  # Prediksi category
    df["predicted_priority"] = priority_model.predict(df["description"])  # Prediksi priority
    df["predicted_category_match"] = df["predicted_category"] == df["category"]  # Cek kecocokan category
    df["predicted_priority_match"] = df["predicted_priority"] == df["priority"]  # Cek kecocokan priority

    category_metrics = (  # DataFrame metrik category
        pd.DataFrame(category_report_dict)  # Ubah laporan ke DataFrame
        .T.loc[:, ["precision", "recall", "f1-score", "support"]]  # Pilih kolom metrik
        .rename_axis("Category Name")  # Set nama index
    )
    priority_metrics = (  # DataFrame metrik priority
        pd.DataFrame(priority_report_dict)  # Ubah laporan ke DataFrame
        .T.loc[:, ["precision", "recall", "f1-score", "support"]]  # Pilih kolom metrik
        .rename_axis("Priority Name")  # Set nama index
    )

    prediction_notes_map = {  # Deskripsi header untuk sheet Predictions
        "subject": "Judul ringkas tiket dari pengguna atau sistem",  # Keterangan subject
        "description": "Detail masalah/kebutuhan; input utama model untuk prediksi",  # Keterangan description
        "answer": "Jawaban/solusi yang diberikan pada tiket",  # Keterangan answer
        "type": "Tipe/klasifikasi tiket dari sumber data",  # Keterangan type
        "priority": "Label priority asli dari dataset (ground truth)",  # Keterangan priority
        "category": "Label category asli dari dataset (ground truth)",  # Keterangan category
        "predicted_category": "Hasil prediksi category berbasis teks description",  # Keterangan predicted_category
        "predicted_priority": "Hasil prediksi priority berbasis teks description",  # Keterangan predicted_priority
        "predicted_category_match": "True jika predicted_category sama dengan category",  # Keterangan match category
        "predicted_priority_match": "True jika predicted_priority sama dengan priority",  # Keterangan match priority
    }
    prediction_notes = {  # Mapping kolom ke komentar Excel
        index + 1: prediction_notes_map.get(col, "")  # Nomor kolom Excel dan note
        for index, col in enumerate(df.columns)  # Iterasi seluruh kolom
    }
    metric_notes = {  # Deskripsi header untuk sheet metrik
        1: "Nama kelas/label yang dievaluasi",  # Kolom nama kelas
        2: "Presisi per kelas: TP / (TP + FP)",  # Kolom precision
        3: "Recall per kelas: TP / (TP + FN)",  # Kolom recall
        4: "F1-score per kelas: 2 * (precision * recall) / (precision + recall)",  # Kolom f1-score
        5: "Jumlah data per kelas pada set evaluasi",  # Kolom support
    }

    def write_output(file_path):  # Fungsi menulis file Excel
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:  # Buka writer Excel
            df.to_excel(writer, index=False, sheet_name="Predictions")  # Tulis sheet prediksi
            category_metrics.to_excel(writer, sheet_name="Category Accuracy")  # Tulis sheet akurasi category
            priority_metrics.to_excel(writer, sheet_name="Priority Accuracy")  # Tulis sheet akurasi priority

            prediction_sheet = writer.sheets["Predictions"]  # Ambil sheet Predictions
            for column_index, note in prediction_notes.items():  # Loop note per kolom
                if note:  # Hanya tulis jika ada note
                    cell = prediction_sheet.cell(row=1, column=column_index)  # Ambil cell header
                    cell.comment = Comment(note, "SVM")  # Tambahkan komentar Excel

            category_sheet = writer.sheets["Category Accuracy"]  # Ambil sheet Category Accuracy
            for column_index, note in metric_notes.items():  # Loop note metrik
                cell = category_sheet.cell(row=1, column=column_index)  # Ambil cell header
                cell.comment = Comment(note, "SVM")  # Tambahkan komentar Excel

            priority_sheet = writer.sheets["Priority Accuracy"]  # Ambil sheet Priority Accuracy
            for column_index, note in metric_notes.items():  # Loop note metrik
                cell = priority_sheet.cell(row=1, column=column_index)  # Ambil cell header
                cell.comment = Comment(note, "SVM")  # Tambahkan komentar Excel

    output_path = "cobacek_pred.xlsx"  # Path output utama
    fallback_path = "cobacek_pred_new.xlsx"  # Path cadangan jika file terkunci
    try:  # Coba tulis ke file utama
        write_output(output_path)  # Tulis output utama
    except PermissionError:  # Tangani file terkunci
        write_output(fallback_path)  # Tulis ke file cadangan


if __name__ == "__main__":  # Entry point script
    main()  # Panggil fungsi utama
